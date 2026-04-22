"""考试核心视图：试题/试卷管理、答题、提交、批阅、成绩查看、错题集。"""
import json
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from classes.models import ClassMember, ClassRoom, Course
from notifications.models import notify
from users.decorators import role_required, student_required, teacher_required
from users.models import UserProfile

from .forms import ExamPaperForm, QuestionForm
from .models import AnswerRecord, ExamPaper, ExamRecord, Question, WrongQuestion


# ============================================================
# 通用工具
# ============================================================
def _get_role(user):
    profile = getattr(user, 'profile', None)
    return getattr(profile, 'role', None)


def _objective_is_correct(question: Question, student_answer: str) -> bool:
    """客观题是否答对。"""
    if not student_answer:
        return False
    stu = (student_answer or '').strip()
    std = (question.answer or '').strip()
    if question.type == 'multiple_choice':
        a = sorted(x.strip().upper() for x in stu.split(',') if x.strip())
        b = sorted(x.strip().upper() for x in std.split(',') if x.strip())
        return a == b
    if question.type == 'single_choice':
        return stu.upper() == std.upper()
    if question.type == 'judgment':
        return stu == std
    return False


def _parse_options(raw: str):
    """把 options 字段文本解析为 [(key, label)] 列表。"""
    items = []
    if not raw:
        return items
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        # 支持 "A.内容" / "A、内容" / "A 内容"
        for sep in ('.', '、', ' '):
            if sep in line:
                key, value = line.split(sep, 1)
                items.append((key.strip(), value.strip()))
                break
        else:
            items.append((line, line))
    return items


def _grade_answer(question: Question, answer_text: str, record: ExamRecord):
    """对一道题自动判分，返回 (is_correct, similarity, score_or_None)。

    客观题返回确定分值；主观题仅给建议分(auto_score)，score 暂留空等教师确认。
    """
    if question.type in Question.OBJECTIVE_TYPES:
        correct = _objective_is_correct(question, answer_text)
        return correct, None, (question.score if correct else 0.0), (question.score if correct else 0.0)

    # 主观题：计算相似度 + 关键词命中，生成建议分
    from scoring.utils import subjective_auto_score
    similarity, auto_score = subjective_auto_score(question, answer_text)
    # 主观题 score 暂时填 auto_score，教师可修改
    return None, similarity, None, auto_score


def _auto_grade_record(record: ExamRecord):
    """批改整张试卷。"""
    objective_total = 0.0
    subjective_auto_total = 0.0
    for ans in record.answers.select_related('question').all():
        q = ans.question
        is_correct, sim, score, auto_score = _grade_answer(q, ans.student_answer or '', record)
        ans.is_correct = is_correct
        ans.similarity = sim
        ans.auto_score = auto_score
        if q.type in Question.OBJECTIVE_TYPES:
            ans.score = score  # 客观题立即给分
            objective_total += score or 0.0
        else:
            # 主观题先不最终给分，留 auto_score 供教师参考
            subjective_auto_total += auto_score or 0.0
        ans.save()

    record.objective_score = round(objective_total, 2)
    # 如果本卷没有主观题，则直接变为 graded，总分=客观分
    has_subjective = record.answers.filter(
        question__type__in=Question.SUBJECTIVE_TYPES
    ).exists()
    if has_subjective:
        record.subjective_score = 0.0  # 等教师确认
        record.score = round(objective_total, 2)  # 暂记客观分
        record.status = 'finished'
    else:
        record.subjective_score = 0.0
        record.score = round(objective_total, 2)
        record.status = 'graded'
    record.save()
    return objective_total, subjective_auto_total


def _sync_wrong_questions(record: ExamRecord):
    """将本次考试中未得满分的题目同步到错题本。"""
    for ans in record.answers.select_related('question').all():
        q = ans.question
        got = ans.score if ans.score is not None else (ans.auto_score or 0)
        if got < q.score:  # 没拿满分就算错题
            WrongQuestion.objects.get_or_create(
                student=record.student, question=q,
                defaults={'answer_record': ans},
            )


# ============================================================
# 仪表盘
# ============================================================
@login_required
def dashboard(request):
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=request.user)

    role = user_profile.role

    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = now - timezone.timedelta(days=7)

    if role == 'teacher':
        question_count = Question.objects.filter(created_by=request.user).count()
        paper_count = ExamPaper.objects.filter(created_by=request.user).count()
        pending_records = ExamRecord.objects.filter(
            paper__created_by=request.user,
            status='finished',
        ).count()
        teaching_courses = Course.objects.filter(teacher=request.user).count()
        headed_classes = request.user.headed_classes.count()
        # 新增：本周提交、平均分
        week_submissions = ExamRecord.objects.filter(
            paper__created_by=request.user,
            end_time__gte=week_ago,
        ).count()
        graded_scores = ExamRecord.objects.filter(
            paper__created_by=request.user, status='graded', score__isnull=False,
        ).values_list('score', flat=True)
        avg_score = round(sum(graded_scores) / len(graded_scores), 1) if graded_scores else 0

        context = {
            'is_teacher': True, 'role': role,
            'question_count': question_count,
            'paper_count': paper_count,
            'pending_correction_count': pending_records,
            'teaching_courses': teaching_courses,
            'headed_classes': headed_classes,
            'week_submission_count': week_submissions,
            'avg_score': avg_score,
            'recent_papers': ExamPaper.objects.filter(
                created_by=request.user).order_by('-created_at')[:5],
            'recent_submissions': ExamRecord.objects.filter(
                paper__created_by=request.user,
                end_time__isnull=False,
            ).select_related('student', 'paper').order_by('-end_time')[:6],
        }

    elif role == 'admin':
        from notifications.models import Notification
        total_user = User.objects.count()
        teacher_count = UserProfile.objects.filter(role='teacher').count()
        student_count = UserProfile.objects.filter(role='student').count()
        admin_count = UserProfile.objects.filter(role='admin').count()
        context = {
            'is_admin': True, 'role': role,
            'total_user': total_user,
            'admin_count': admin_count,
            'teacher_count': teacher_count,
            'student_count': student_count,
            # 系统整体概况
            'total_class': ClassRoom.objects.count(),
            'total_course': Course.objects.count(),
            'total_paper': ExamPaper.objects.count(),
            'total_paper_published': ExamPaper.objects.filter(is_published=True).count(),
            'total_record': ExamRecord.objects.count(),
            'total_question': Question.objects.count(),
            # 近期动态
            'new_users_week': User.objects.filter(date_joined__gte=week_ago).count(),
            'new_papers_week': ExamPaper.objects.filter(created_at__gte=week_ago).count(),
            'new_records_today': ExamRecord.objects.filter(
                start_time__gte=today_start).count(),
            'recent_users': User.objects.select_related('profile').order_by(
                '-date_joined')[:5],
            'recent_papers_sys': ExamPaper.objects.select_related(
                'created_by', 'course').order_by('-created_at')[:5],
            'recent_broadcasts': Notification.objects.filter(
                sender=request.user, type='generic').order_by('-created_at')[:5],
        }

    else:  # student
        class_ids = ClassMember.objects.filter(
            user=request.user, role='student'
        ).values_list('classroom_id', flat=True)
        available_papers = ExamPaper.objects.filter(is_published=True).filter(
            Q(course__isnull=True) | Q(course__classroom_id__in=class_ids)
        )
        participated_ids = ExamRecord.objects.filter(
            student=request.user
        ).values_list('paper_id', flat=True)
        uncompleted = available_papers.exclude(id__in=participated_ids).distinct()
        completed_qs = ExamRecord.objects.filter(
            student=request.user, status__in=['finished', 'graded'],
        ).select_related('paper').order_by('-end_time')
        joined_class_count = ClassMember.objects.filter(
            user=request.user, role='student'
        ).count()
        wrong_count = WrongQuestion.objects.filter(student=request.user).count()

        graded_scores = list(completed_qs.filter(
            status='graded', score__isnull=False,
        ).values_list('score', flat=True))
        avg_score = round(sum(graded_scores) / len(graded_scores), 1) if graded_scores else 0

        context = {
            'is_teacher': False, 'role': role,
            'completed_paper_count': completed_qs.count(),
            'uncompleted_paper_count': uncompleted.count(),
            'joined_class_count': joined_class_count,
            'wrong_count': wrong_count,
            'avg_score': avg_score,
            'upcoming_papers': uncompleted.order_by('-created_at')[:5],
            'recent_results': completed_qs[:5],
            'my_classes': ClassRoom.objects.filter(
                id__in=list(class_ids),
            ).select_related('head_teacher')[:6],
        }

    return render(request, 'exams/dashboard.html', context)


# ============================================================
# 学生端
# ============================================================
@student_required
def student_paper_list(request):
    """学生：我可以参加的试卷（未完成）。"""
    class_ids = ClassMember.objects.filter(
        user=request.user, role='student',
    ).values_list('classroom_id', flat=True)
    papers = ExamPaper.objects.filter(is_published=True).filter(
        Q(course__isnull=True) | Q(course__classroom_id__in=class_ids)
    ).distinct()

    participated_ids = ExamRecord.objects.filter(
        student=request.user, status__in=['finished', 'graded']
    ).values_list('paper_id', flat=True)
    uncompleted = papers.exclude(id__in=participated_ids)

    context = {
        'uncompleted_papers': uncompleted,
        'paper_count': uncompleted.count(),
    }
    return render(request, 'exams/student_paper_list.html', context)


@student_required
def my_todo(request):
    """学生"我的待办"：按截止时间排序，显示剩余天数并给出紧急标识。

    规则：
      - 已发布、且学生有权访问的试卷
      - 尚未提交（unfinished 或无记录）
      - 有 deadline 的按时间升序；无 deadline 的排最后
    """
    class_ids = ClassMember.objects.filter(
        user=request.user, role='student',
    ).values_list('classroom_id', flat=True)
    papers = ExamPaper.objects.filter(is_published=True).filter(
        Q(course__isnull=True) | Q(course__classroom_id__in=class_ids)
    ).distinct().select_related('course', 'course__classroom')

    participated_ids = ExamRecord.objects.filter(
        student=request.user, status__in=['finished', 'graded']
    ).values_list('paper_id', flat=True)
    pending = papers.exclude(id__in=participated_ids)

    now = timezone.now()
    todo_rows = []
    for p in pending:
        if p.deadline:
            delta = p.deadline - now
            total_sec = int(delta.total_seconds())
            if total_sec < 0:
                urgency = 'overdue'
                hint = '已截止'
            elif total_sec < 86400:  # <1 天
                urgency = 'urgent'
                hint = f'剩余 {total_sec // 3600} 小时'
            elif total_sec < 86400 * 3:
                urgency = 'soon'
                hint = f'剩余 {total_sec // 86400} 天'
            else:
                urgency = 'normal'
                hint = f'剩余 {total_sec // 86400} 天'
        else:
            urgency = 'normal'
            hint = '长期有效'
        todo_rows.append({
            'paper': p, 'urgency': urgency, 'hint': hint,
        })

    # 按 urgency 优先级 + deadline 升序（无 deadline 的排最后）
    import datetime as _dt
    far_future = _dt.datetime(9999, 12, 31, tzinfo=now.tzinfo)
    order = {'overdue': 0, 'urgent': 1, 'soon': 2, 'normal': 3}
    todo_rows.sort(key=lambda r: (
        order[r['urgency']],
        r['paper'].deadline or far_future,
    ))

    return render(request, 'exams/my_todo.html', {
        'todo_rows': todo_rows,
        'todo_count': len(todo_rows),
    })


@student_required
def exam_paper_detail(request, paper_id):
    """学生答题界面。

    实现 "答题草稿 + 续时" 两大逻辑:
      - 已作答过的题目答案会自动回填到输入框
      - 倒计时按 `start_time` + `duration` 计算真实剩余秒数
        (中途退出再进入,时间不会被重置)
      - 若已超时则自动提交
    """
    paper = get_object_or_404(ExamPaper, id=paper_id, is_published=True)
    if not paper.is_accessible_to(request.user):
        messages.error(request, '你不在该试卷所属班级，无法作答')
        return redirect('exams:student_paper_list')

    record, created = ExamRecord.objects.get_or_create(
        student=request.user, paper=paper
    )
    if record.status in ('finished', 'graded'):
        messages.info(request, '该试卷已提交，跳转到成绩详情')
        return redirect('exams:exam_result', record_id=record.id)

    # ---------- 续时:真实剩余秒数 ----------
    elapsed = (timezone.now() - record.start_time).total_seconds()
    remaining_seconds = max(0, int(paper.duration * 60 - elapsed))
    if remaining_seconds <= 0:
        # 超时自动提交(仅以已有草稿作答)
        messages.warning(request, '考试时间已到,系统已为你自动提交。')
        return redirect('exams:submit_exam', paper_id=paper.id)

    # 预处理分组
    grouped = defaultdict(list)
    type_order = ['single_choice', 'multiple_choice', 'judgment', 'short_answer']
    type_cn = dict(Question.TYPE_CHOICES)

    for q in paper.questions.all():
        item = {
            'id': q.id,
            'content': q.content,
            'type': q.type,
            'score': q.score,
            'options': [],
        }
        if q.type in ('single_choice', 'multiple_choice'):
            opts = _parse_options(q.options or '')
            if not opts:
                opts = [('A', '选项A'), ('B', '选项B'), ('C', '选项C'), ('D', '选项D')]
            item['options'] = opts
        elif q.type == 'judgment':
            item['options'] = [('对', '对'), ('错', '错')]
        grouped[q.type].append(item)

    question_groups = []
    roman = ['一', '二', '三', '四', '五']
    for i, t in enumerate(type_order):
        if grouped.get(t):
            question_groups.append((f'{roman[i]}、{type_cn[t]}', grouped[t]))

    # 恢复已填过的答案（未提交时）
    saved_answers = {
        str(a.question_id): a.student_answer or ''
        for a in record.answers.all()
    }

    context = {
        'paper': paper,
        'question_groups': question_groups,
        'duration': paper.duration,
        'remaining_seconds': remaining_seconds,
        'record_id': record.id,
        'saved_answers': saved_answers,
    }
    return render(request, 'exams/exam_paper_detail.html', context)


@student_required
@require_http_methods(['POST'])
def autosave_answer(request, paper_id):
    """每道题的答案草稿自动保存(AJAX 调用)。

    请求体:
      - field = 'q_<qid>'
      - value = 'A'              (单选/判断)
      - value = 'A,B'            (多选,逗号分隔)
      - value = '长文本...'       (简答)
    响应:{"ok": true, "saved_at": <ISO>}
    """
    from django.http import JsonResponse
    paper = get_object_or_404(ExamPaper, id=paper_id, is_published=True)
    record, _ = ExamRecord.objects.get_or_create(
        student=request.user, paper=paper)
    if record.status in ('finished', 'graded'):
        return JsonResponse({'ok': False, 'reason': 'already_submitted'}, status=400)

    # 时间过期后拒绝保存
    elapsed = (timezone.now() - record.start_time).total_seconds()
    if elapsed > paper.duration * 60:
        return JsonResponse({'ok': False, 'reason': 'timeout'}, status=400)

    field = request.POST.get('field', '')
    value = request.POST.get('value', '')
    if not field.startswith('q_'):
        return JsonResponse({'ok': False, 'reason': 'bad_field'}, status=400)
    try:
        qid = int(field[2:])
    except ValueError:
        return JsonResponse({'ok': False, 'reason': 'bad_field'}, status=400)

    # 校验题目确实属于本卷
    if not paper.questions.filter(id=qid).exists():
        return JsonResponse({'ok': False, 'reason': 'bad_question'}, status=400)

    AnswerRecord.objects.update_or_create(
        record=record, question_id=qid,
        defaults={'student_answer': value},
    )
    return JsonResponse({'ok': True, 'saved_at': timezone.now().isoformat()})


@student_required
@transaction.atomic
def submit_exam(request, paper_id):
    """学生提交试卷：自动评分客观题 → 跳到结果页。"""
    paper = get_object_or_404(ExamPaper, id=paper_id, is_published=True)
    record, _ = ExamRecord.objects.get_or_create(student=request.user, paper=paper)
    if record.status in ('finished', 'graded'):
        messages.error(request, '该试卷已提交，不可重复提交')
        return redirect('exams:exam_result', record_id=record.id)

    if request.method != 'POST':
        return redirect('exams:exam_paper_detail', paper_id=paper_id)

    # 保存每题作答
    for q in paper.questions.all():
        if q.type == 'multiple_choice':
            answer_text = ','.join(sorted(request.POST.getlist(f'q_{q.id}')))
        else:
            answer_text = request.POST.get(f'q_{q.id}', '').strip()
        AnswerRecord.objects.update_or_create(
            record=record, question=q,
            defaults={'student_answer': answer_text},
        )

    # 自动批改（客观题立即出分）
    _auto_grade_record(record)
    record.end_time = timezone.now()
    record.save()

    # 错题本同步
    _sync_wrong_questions(record)

    # 消息通知教师
    notify(
        recipient=paper.created_by,
        sender=request.user,
        type='exam_submitted',
        title=f'学生 {request.user.username} 提交了《{paper.title}》',
        message=f'客观题得分：{record.objective_score}分，共{record.answers.count()}题',
        link=f'/exams/teacher/correction/{record.id}/',
    )

    # 如果没有主观题，直接给学生提醒成绩已出
    if record.status == 'graded':
        notify(
            recipient=request.user, sender=None,
            type='exam_graded',
            title=f'《{paper.title}》成绩已出',
            message=f'你的总得分：{record.score}分',
            link=f'/exams/result/{record.id}/',
        )
        messages.success(request, f'试卷提交成功！你的得分：{record.score}分')
    else:
        messages.success(
            request,
            f'试卷已提交，客观题得分 {record.objective_score} 分，'
            '主观题待教师批阅。',
        )
    return redirect('exams:exam_result', record_id=record.id)


@login_required
def exam_result(request, record_id):
    """考试详情：客观题对错逐题展示、主观题显示相似度与教师评分。"""
    record = get_object_or_404(
        ExamRecord.objects.select_related('paper', 'student'), id=record_id,
    )
    # 权限：本人/试卷创建者/班级任课教师/管理员
    role = _get_role(request.user)
    allowed = (
        record.student_id == request.user.id
        or record.paper.created_by_id == request.user.id
        or role == 'admin'
    )
    if not allowed and role == 'teacher':
        # 检查该教师是否在试卷所属班级任教
        course = record.paper.course
        if course is not None:
            allowed = ClassMember.objects.filter(
                classroom=course.classroom, user=request.user, role='teacher'
            ).exists()
    if not allowed:
        messages.error(request, '无权查看该考试')
        return redirect('exams:dashboard')

    answers = record.answers.select_related('question').all()

    # 按题型分组
    grouped = defaultdict(list)
    type_cn = dict(Question.TYPE_CHOICES)
    for a in answers:
        grouped[a.question.type].append(a)

    type_order = ['single_choice', 'multiple_choice', 'judgment', 'short_answer']
    result_groups = []
    roman = ['一', '二', '三', '四']
    for i, t in enumerate(type_order):
        rows = grouped.get(t) or []
        if not rows:
            continue
        result_groups.append((f'{roman[i]}、{type_cn[t]}', t, rows))

    context = {
        'record': record,
        'paper': record.paper,
        'result_groups': result_groups,
        'is_student': role == 'student',
    }
    return render(request, 'exams/exam_result.html', context)


@student_required
def student_record_list(request):
    """学生：我的考试记录。"""
    records = ExamRecord.objects.filter(
        student=request.user
    ).select_related('paper').order_by('-start_time')
    return render(request, 'exams/student_record_list.html', {
        'exam_records': records,
    })


# ============================================================
# 学生：错题集
# ============================================================
@student_required
def wrong_question_list(request):
    subject = request.GET.get('subject', '').strip()
    qtype = request.GET.get('type', '').strip()
    favorite_only = request.GET.get('favorite') == '1'

    qs = WrongQuestion.objects.filter(student=request.user).select_related(
        'question', 'answer_record'
    )
    if subject:
        qs = qs.filter(question__subject=subject)
    if qtype:
        qs = qs.filter(question__type=qtype)
    if favorite_only:
        qs = qs.filter(is_favorite=True)

    # 取得已有学科/题型作为筛选选项
    subjects = Question.objects.filter(
        wrongquestion__student=request.user,
    ).values_list('subject', flat=True).distinct()
    subjects = [s for s in subjects if s]

    return render(request, 'exams/wrong_question_list.html', {
        'wrongs': qs,
        'subjects': subjects,
        'type_choices': Question.TYPE_CHOICES,
        'current_subject': subject,
        'current_type': qtype,
        'favorite_only': favorite_only,
    })


@student_required
def wrong_question_toggle(request, wq_id):
    """收藏/取消收藏错题。"""
    wq = get_object_or_404(WrongQuestion, id=wq_id, student=request.user)
    wq.is_favorite = not wq.is_favorite
    wq.save(update_fields=['is_favorite'])
    messages.success(request, '已收藏' if wq.is_favorite else '已取消收藏')
    return redirect('exams:wrong_question_list')


@student_required
def wrong_question_delete(request, wq_id):
    wq = get_object_or_404(WrongQuestion, id=wq_id, student=request.user)
    wq.delete()
    messages.success(request, '已从错题集移除')
    return redirect('exams:wrong_question_list')


# ============================================================
# 教师端：试题管理
# ============================================================
@teacher_required
def question_list(request):
    qs = Question.objects.filter(created_by=request.user)
    subject = request.GET.get('subject', '').strip()
    qtype = request.GET.get('type', '').strip()
    if subject:
        qs = qs.filter(subject=subject)
    if qtype:
        qs = qs.filter(type=qtype)

    subjects = Question.objects.filter(
        created_by=request.user
    ).values_list('subject', flat=True).distinct()
    subjects = [s for s in subjects if s]

    return render(request, 'exams/question_list.html', {
        'questions': qs,
        'subjects': subjects,
        'type_choices': Question.TYPE_CHOICES,
        'current_subject': subject,
        'current_type': qtype,
    })


@teacher_required
def question_create(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            q = form.save(commit=False)
            q.created_by = request.user
            q.save()
            messages.success(request, '试题创建成功')
            return redirect('exams:question_list')
    else:
        form = QuestionForm()
    return render(request, 'exams/question_form.html', {
        'form': form, 'title': '添加试题',
    })


@teacher_required
def question_edit(request, qid):
    q = get_object_or_404(Question, id=qid, created_by=request.user)
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=q)
        if form.is_valid():
            form.save()
            messages.success(request, '试题已更新')
            return redirect('exams:question_list')
    else:
        form = QuestionForm(instance=q)
    return render(request, 'exams/question_form.html', {
        'form': form, 'title': '编辑试题', 'question': q,
    })


@teacher_required
def question_delete(request, qid):
    q = get_object_or_404(Question, id=qid, created_by=request.user)
    q.delete()
    messages.success(request, '试题已删除')
    return redirect('exams:question_list')


# ============================================================
# 教师端：试卷管理
# ============================================================
@teacher_required
def paper_list(request):
    papers = ExamPaper.objects.filter(created_by=request.user).select_related('course', 'course__classroom')
    return render(request, 'exams/paper_list.html', {'papers': papers})


@teacher_required
def paper_create(request):
    if request.method == 'POST':
        form = ExamPaperForm(request.POST, teacher=request.user)
        if form.is_valid():
            paper = form.save(commit=False)
            paper.created_by = request.user
            paper.save()
            form.save_m2m()

            # 若已发布 + 绑定课程，通知班上所有学生
            if paper.is_published and paper.course_id:
                students = User.objects.filter(
                    class_memberships__classroom_id=paper.course.classroom_id,
                    class_memberships__role='student',
                )
                for s in students:
                    notify(
                        recipient=s, sender=request.user,
                        type='exam_published',
                        title=f'新试卷《{paper.title}》已发布',
                        message=f'课程：{paper.course.name}，时长 {paper.duration} 分钟',
                        link=f'/exams/student/paper/{paper.id}/',
                    )

            messages.success(request, '试卷创建成功')
            return redirect('exams:paper_list')
    else:
        form = ExamPaperForm(teacher=request.user)
    return render(request, 'exams/paper_form.html', {
        'form': form, 'title': '创建试卷',
    })


@teacher_required
def paper_edit(request, paper_id):
    paper = get_object_or_404(ExamPaper, id=paper_id, created_by=request.user)
    prev_published = paper.is_published
    if request.method == 'POST':
        form = ExamPaperForm(request.POST, instance=paper, teacher=request.user)
        if form.is_valid():
            p = form.save()
            # 从未发布 → 已发布 也要通知
            if not prev_published and p.is_published and p.course_id:
                students = User.objects.filter(
                    class_memberships__classroom_id=p.course.classroom_id,
                    class_memberships__role='student',
                )
                for s in students:
                    notify(
                        recipient=s, sender=request.user,
                        type='exam_published',
                        title=f'新试卷《{p.title}》已发布',
                        link=f'/exams/student/paper/{p.id}/',
                    )
            messages.success(request, '试卷更新成功')
            return redirect('exams:paper_list')
    else:
        form = ExamPaperForm(instance=paper, teacher=request.user)
    return render(request, 'exams/paper_form.html', {
        'form': form, 'title': '编辑试卷', 'paper': paper,
    })


@teacher_required
def paper_delete(request, paper_id):
    p = get_object_or_404(ExamPaper, id=paper_id, created_by=request.user)
    p.delete()
    messages.success(request, '试卷已删除')
    return redirect('exams:paper_list')


# ============================================================
# 教师端：阅卷
# ============================================================
@teacher_required
def teacher_correction(request):
    """教师的阅卷首页：待批阅列表。"""
    teacher_papers = ExamPaper.objects.filter(created_by=request.user)
    records = ExamRecord.objects.filter(
        paper__in=teacher_papers,
    ).select_related('student', 'paper').order_by('-end_time')

    # 过滤：只看待批阅 / 全部
    show = request.GET.get('filter', 'pending')
    if show == 'pending':
        records = records.filter(status='finished')
    elif show == 'graded':
        records = records.filter(status='graded')

    return render(request, 'exams/teacher_correction.html', {
        'correction_records': records,
        'filter': show,
    })


@teacher_required
def export_records_xlsx(request):
    """教师导出自己试卷的全部考试记录为 Excel。

    可选 query 参数:
      - paper_id: 仅导出指定试卷
      - status:   finished / graded / all
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    teacher_papers = ExamPaper.objects.filter(created_by=request.user)
    records = ExamRecord.objects.filter(
        paper__in=teacher_papers,
        end_time__isnull=False,
    ).select_related('student', 'paper', 'student__profile').order_by(
        'paper_id', '-end_time')

    paper_id = request.GET.get('paper_id')
    if paper_id:
        records = records.filter(paper_id=paper_id)
    status = request.GET.get('status', 'all')
    if status in ('finished', 'graded'):
        records = records.filter(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = '成绩单'

    headers = ['考生工/学号', '考生姓名', '用户名', '试卷标题', '提交时间',
               '客观分', '主观分', '总分', '状态']
    ws.append(headers)
    # 表头样式
    head_fill = PatternFill('solid', fgColor='C42A2A')
    head_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center

    status_display = {
        'unfinished': '未完成', 'finished': '待批阅', 'graded': '已批阅',
    }
    for r in records:
        ws.append([
            getattr(r.student.profile, 'student_id', '') or '',
            r.student.first_name or '',
            r.student.username,
            r.paper.title,
            r.end_time.strftime('%Y-%m-%d %H:%M') if r.end_time else '',
            r.objective_score or 0,
            r.subjective_score or 0,
            r.score if r.score is not None else '',
            status_display.get(r.status, r.status),
        ])

    # 自动列宽
    for col in ws.columns:
        max_len = 10
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) + 2)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len, 40)

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    fname = quote(f'成绩单_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{fname}"
    return resp


@teacher_required
def paper_stats(request, paper_id):
    """教师查看某份试卷的成绩统计(分数分布 + 及格率)。"""
    paper = get_object_or_404(ExamPaper, id=paper_id, created_by=request.user)
    records = ExamRecord.objects.filter(
        paper=paper, status='graded', score__isnull=False,
    ).select_related('student').order_by('-score')

    total_full = paper.total_score  # property 计算满分
    scores = [r.score for r in records]
    n = len(scores)

    # 分数分布 5 档 (60 分制映射到 total_full)
    bins = [0, 0, 0, 0, 0]  # <60, 60-69, 70-79, 80-89, 90-100
    bin_labels = ['不及格', '60-69', '70-79', '80-89', '90-100']
    for s in scores:
        # 归一化到 100 制方便统计
        pct = (s / total_full * 100) if total_full else 0
        if pct < 60:
            bins[0] += 1
        elif pct < 70:
            bins[1] += 1
        elif pct < 80:
            bins[2] += 1
        elif pct < 90:
            bins[3] += 1
        else:
            bins[4] += 1

    pass_count = sum(bins[1:])
    fail_count = bins[0]
    pass_rate = round(pass_count / n * 100, 1) if n else 0
    avg = round(sum(scores) / n, 2) if n else 0
    highest = max(scores) if scores else 0
    lowest = min(scores) if scores else 0

    return render(request, 'exams/paper_stats.html', {
        'paper': paper, 'records': records,
        'total_full': total_full,
        'n': n, 'avg': avg, 'highest': highest, 'lowest': lowest,
        'pass_count': pass_count, 'fail_count': fail_count,
        'pass_rate': pass_rate,
        'bin_labels': bin_labels, 'bins': bins,
    })


@teacher_required
@transaction.atomic
def grade_record(request, record_id):
    """对某份试卷进行批阅（主要为主观题打分与复核）。"""
    record = get_object_or_404(
        ExamRecord.objects.select_related('paper', 'student'),
        id=record_id, paper__created_by=request.user,
    )
    answers = list(record.answers.select_related('question').all())

    if request.method == 'POST':
        # 根据前端提交的分数覆盖 score
        objective_total = 0.0
        subjective_total = 0.0
        for a in answers:
            q = a.question
            key = f'score_{a.id}'
            raw = request.POST.get(key)
            if raw is not None and raw != '':
                try:
                    value = max(0.0, min(float(raw), q.score))
                except ValueError:
                    value = 0.0
                a.score = value
                a.is_correct = (value >= q.score)
                a.save(update_fields=['score', 'is_correct'])
            if q.type in Question.OBJECTIVE_TYPES:
                objective_total += a.score or 0
            else:
                subjective_total += a.score or 0

        record.objective_score = round(objective_total, 2)
        record.subjective_score = round(subjective_total, 2)
        record.score = round(objective_total + subjective_total, 2)
        record.status = 'graded'
        record.save()

        # 同步错题本（根据最终得分）
        _sync_wrong_questions(record)

        # 通知学生
        notify(
            recipient=record.student, sender=request.user,
            type='exam_graded',
            title=f'《{record.paper.title}》批阅完成',
            message=f'你的总分：{record.score}分',
            link=f'/exams/result/{record.id}/',
        )
        # 给教师自己一条完成提醒
        notify(
            recipient=request.user, sender=None,
            type='grade_finished',
            title=f'你已完成《{record.paper.title}》的批阅',
            message=f'学生：{record.student.username}',
            link=f'/exams/result/{record.id}/',
        )
        messages.success(request, '批阅完成！')
        return redirect('exams:teacher_correction')

    return render(request, 'exams/grade_record.html', {
        'record': record,
        'answers': answers,
    })
